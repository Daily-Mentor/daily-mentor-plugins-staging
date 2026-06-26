"""File discovery and parsing. No business logic — just typed frames.

The 7-file pack:
    shopify_daily_sales_*.csv
    Gross sales by new or returning customer - *.csv
    Sessions by month - *.csv
    *_Profit_and_Loss.xlsx           (Xero)
    *_Balance_Sheet.xlsx             (Xero)
    *_Account_Transactions.xlsx      (Xero)
    facebook_spend.{csv,xlsx}        (optional: google_spend, tiktok_spend)
    Optional: shopify cohort export

Currency is sniffed from CSV column headers (e.g. `Amount spent (NZD)`).
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .models import CurrencyTag, IngestBundle, IngestMeta


# ---------- File discovery ----------


def _scan(inputs_dir: Path) -> dict[str, Path]:
    """Map role → file path. Tolerant of naming variations.

    Multi-entity packs (e.g. parent + subsidiary Xero exports) — pick the
    largest file for each Xero role. Larger = more rows = more coverage =
    the operating entity. Other entities' files are dropped from the build.
    """
    files = {p.name: p for p in inputs_dir.iterdir() if p.is_file()}
    roles: dict[str, Path] = {}
    candidate_atxn: list[Path] = []
    candidate_pl: list[Path] = []
    candidate_bs: list[Path] = []
    candidate_shopify_daily: list[Path] = []
    for name, path in files.items():
        lower = name.lower()
        # Daily sales: 'Daily Mentor - Total Sales Over Time' (new) or legacy 'shopify_daily' / 'total_sales'.
        if (("total sales over time" in lower or "shopify_daily" in lower or "total_sales" in lower)
                and lower.endswith(".csv")):
            candidate_shopify_daily.append(path)
        # NC vs RC: 'Daily Mentor - NC v RC L365' (new) or legacy 'gross sales by new or returning'.
        elif ("nc v rc" in lower or "nc vs rc" in lower or "new or returning" in lower
              or "new vs returning" in lower):
            roles["nc_rc"] = path
        elif "sessions by month" in lower or ("sessions" in lower and lower.endswith(".csv")):
            roles["sessions"] = path
        elif "profit_and_loss" in lower and lower.endswith(".xlsx"):
            candidate_pl.append(path)
        elif "balance_sheet" in lower and lower.endswith(".xlsx"):
            candidate_bs.append(path)
        elif "account_transactions" in lower and lower.endswith(".xlsx"):
            candidate_atxn.append(path)
        elif "facebook" in lower and "spend" in lower:
            roles.setdefault("ad_spend_meta", path)
        elif "google" in lower and "spend" in lower:
            roles["ad_spend_google"] = path
        elif "tiktok" in lower and "spend" in lower:
            roles["ad_spend_tiktok"] = path
        elif "cohort" in lower:
            roles["cohort"] = path
    if candidate_shopify_daily:
        # Multiple 'total sales over time' exports (e.g. a daily file plus a stray
        # monthly one) — prefer the largest, which is the daily-granularity file the
        # P&L needs. Mirrors the pre-flight's largest-wins rule so they never disagree.
        roles["shopify_daily"] = max(candidate_shopify_daily, key=lambda p: p.stat().st_size)
    if candidate_atxn:
        roles["xero_atxn"] = max(candidate_atxn, key=lambda p: p.stat().st_size)
    if candidate_pl:
        roles["xero_pl"] = max(candidate_pl, key=lambda p: p.stat().st_size)
    if candidate_bs:
        roles["xero_bs"] = max(candidate_bs, key=lambda p: p.stat().st_size)
    return roles


def _client_name_from_files(roles: dict[str, Path]) -> str:
    """Infer client name from a Xero export filename like '<Client>_-_Profit_and_Loss.xlsx'."""
    for k in ("xero_pl", "xero_bs", "xero_atxn"):
        if k in roles:
            stem = roles[k].stem  # e.g. "Boosh_PTY_LTD_-_Profit_and_Loss"
            # Strip the report suffix
            for suffix in ("_-_Profit_and_Loss", "_-_Balance_Sheet", "_-_Account_Transactions"):
                if stem.endswith(suffix):
                    return stem[: -len(suffix)].replace("_", " ")
    return "Unknown Client"


# ---------- CSV parsers ----------


def _read_shopify_daily(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # Normalise headers (some exports add leading/trailing whitespace, BOM)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "Day": "day",
        "Orders": "orders",
        "Gross sales": "gross",
        "Discounts": "discounts",
        "Returns": "returns",
        "Net sales": "net",
        "Shipping charges": "shipping",
        "Duties": "duties",
        "Additional fees": "additional_fees",
        "Taxes": "taxes",
        "Total sales": "total",
        "Cost of goods sold": "cogs",
        "Net items sold": "units",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "day" in df.columns:
        df["day"] = pd.to_datetime(df["day"], errors="coerce").dt.date
    for col in ("orders", "gross", "discounts", "returns", "net", "shipping", "taxes", "total", "cogs", "units"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.dropna(subset=["day"]).sort_values("day").reset_index(drop=True)


def _quarter_of(d: date) -> str:
    """Calendar quarter label, e.g. date(2026,2,1) -> 'Q1 2026'."""
    q = (d.month - 1) // 3 + 1
    return f"Q{q} {d.year}"


def _read_nc_rc(path: Path) -> pd.DataFrame:
    """New vs Returning customer split.

    Now period-aware: the L365 export may carry a time dimension (a month, quarter,
    or period-start column) so the NCCM can run quarter-over-quarter. When present,
    each row is tagged with its calendar quarter. When absent (legacy single-aggregate
    export), all rows fall under a single 'All' quarter and downstream code treats the
    whole file as one period.
    """
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "New or returning customer": "segment",
        "New or returning": "segment",
        "Customer type": "segment",
        "Gross sales": "gross",
        "Discounts": "discounts",
        "Returns": "returns",
        "Shipping charges": "shipping",
        "Taxes": "taxes",
        "Orders": "orders",
        "Average order value": "aov",
        "Cost of goods sold": "cogs",
        # possible period columns
        "Month": "period",
        "Quarter": "period",
        "Week": "period",
        "Day": "period",
        "Date": "period",
        "Period": "period",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    for col in ("gross", "discounts", "returns", "shipping", "taxes", "orders", "aov", "cogs"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Derive calendar quarter from a period column if we found one.
    if "period" in df.columns:
        parsed = pd.to_datetime(df["period"], errors="coerce")
        if parsed.notna().any():
            df["period_date"] = parsed.dt.date
            df["quarter"] = parsed.dt.date.map(lambda d: _quarter_of(d) if pd.notna(d) else "All")
        else:
            # period column held quarter labels like 'Q1 2026' already
            df["quarter"] = df["period"].astype(str).str.strip()
    else:
        df["quarter"] = "All"
    return df


_MONTH_OFFSET_RE = re.compile(r"(?:month|period|m)\s*\.?\s*(\d+)", re.IGNORECASE)


def _read_cohort(path: Path) -> pd.DataFrame | None:
    """Normalise a Shopify Cohort Analysis 'Customer value by month' export into a
    tidy frame: columns [cohort, month_offset, value].

    Tolerant of two common layouts:

      WIDE  — one row per acquisition cohort, a column per month-since-acquisition:
              Cohort, Customers, Month 0, Month 1, Month 2, ...
      LONG  — one row per (cohort, month-offset):
              Cohort month, Months since first purchase, Amount spent per customer

    `value` is the cumulative customer value at that month-offset (Shopify reports
    cumulative spend per customer for this report type). Returns None if the file
    can't be coerced into a recognised shape.
    """
    try:
        raw = pd.read_csv(path)
    except Exception:
        return None
    if raw.empty:
        return None
    raw.columns = [str(c).strip().lstrip("﻿") for c in raw.columns]

    # Identify a cohort-label column (first textual/date column that isn't a count).
    cohort_col = None
    for cand in ("Cohort", "Cohort month", "Customer cohort", "Cohort period", "Month", "First order month", "Acquisition month"):
        for col in raw.columns:
            if col.strip().lower() == cand.strip().lower():
                cohort_col = col
                break
        if cohort_col:
            break
    if cohort_col is None:
        cohort_col = raw.columns[0]  # fall back to first column

    # LONG layout: a "months since" column + a single value column.
    months_since_col = None
    for col in raw.columns:
        cl = col.lower()
        if ("since" in cl and ("month" in cl or "purchase" in cl)) or cl in ("months", "period", "month offset"):
            months_since_col = col
            break

    rows: list[dict] = []
    if months_since_col is not None:
        value_col = None
        kind = "value"
        for col in raw.columns:
            cl = col.lower()
            if any(k in cl for k in ("amount", "value", "spent", "revenue", "ltv")):
                value_col = col
                kind = "value"
                break
        if value_col is None:
            # Retention-style cohort (Shopify 'Customer retention rate' export): no $
            # column, but a retention-rate column we can turn into an LTV curve later.
            for col in raw.columns:
                cl = col.lower()
                if "retention" in cl or cl.endswith("rate"):
                    value_col = col
                    kind = "retention"
                    break
        if value_col is None:
            return None
        size_col = None
        for col in raw.columns:
            if "in cohort" in col.lower() or col.lower() in ("cohort size", "customers in cohort"):
                size_col = col
                break
        for _, r in raw.iterrows():
            try:
                off = int(float(str(r[months_since_col]).strip().lower().lstrip("m").strip()))
            except (ValueError, TypeError):
                continue
            val = pd.to_numeric(r[value_col], errors="coerce")
            if pd.isna(val):
                continue
            row = {"cohort": str(r[cohort_col]).strip(), "month_offset": off, "value": float(val), "kind": kind}
            if size_col is not None:
                sz = pd.to_numeric(r[size_col], errors="coerce")
                row["cohort_size"] = float(sz) if not pd.isna(sz) else None
            rows.append(row)
    else:
        # WIDE layout: detect month-offset columns by header pattern.
        offset_cols: list[tuple[str, int]] = []
        for col in raw.columns:
            m = _MONTH_OFFSET_RE.search(col)
            if m and col != cohort_col:
                offset_cols.append((col, int(m.group(1))))
        if not offset_cols:
            return None
        for _, r in raw.iterrows():
            cohort_label = str(r[cohort_col]).strip()
            if not cohort_label or cohort_label.lower() == "nan":
                continue
            for col, off in offset_cols:
                val = pd.to_numeric(r[col], errors="coerce")
                if pd.isna(val):
                    continue
                rows.append({"cohort": cohort_label, "month_offset": off, "value": float(val)})

    if not rows:
        return None
    df = pd.DataFrame(rows).sort_values(["cohort", "month_offset"]).reset_index(drop=True)
    if "kind" not in df.columns:
        df["kind"] = "value"
    return df


def _read_sessions(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "Month": "month",
        "Online store visitors": "visitors",
        "Sessions": "sessions",
        "Conversion rate": "cr",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "month" in df.columns:
        df["month"] = pd.to_datetime(df["month"], errors="coerce").dt.date
    for col in ("visitors", "sessions", "cr"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.dropna(subset=["month"]).sort_values("month").reset_index(drop=True)


_AD_CCY_RE = re.compile(r"Amount spent \(([A-Z]{3})\)")


def _read_ad_spend(path: Path) -> tuple[pd.DataFrame, CurrencyTag]:
    """Read FB/Google/TikTok spend CSV. Returns (df, currency_tag).

    Expected columns: Day, Campaign name, Amount spent (XXX), Reporting starts, Reporting ends.
    """
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    # Find the spend column
    spend_col = None
    ccy = "AUD"
    ccy_confidence = "assumed"
    for col in df.columns:
        m = _AD_CCY_RE.match(col)
        if m:
            spend_col = col
            ccy = m.group(1)
            ccy_confidence = "explicit"
            break
    if spend_col is None:
        # Fallback: look for any 'Amount spent' / 'Spend' / 'Cost' column
        for col in df.columns:
            cl = col.lower()
            if "amount spent" in cl or cl == "spend" or cl == "cost":
                spend_col = col
                break
    if spend_col is None:
        raise ValueError(f"Could not find spend column in {path.name}")
    df = df.rename(columns={spend_col: "amount_orig"})
    rename = {"Day": "day", "Campaign name": "campaign", "Reporting starts": "start", "Reporting ends": "end"}
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "day" in df.columns:
        df["day"] = pd.to_datetime(df["day"], errors="coerce").dt.date
    df["amount_orig"] = pd.to_numeric(df["amount_orig"], errors="coerce").fillna(0)
    df["ccy"] = ccy
    df = df.dropna(subset=["day"]).reset_index(drop=True)
    return df, CurrencyTag(code=ccy, confidence=ccy_confidence, source=spend_col)


# ---------- Xero xlsx parsers ----------


_MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def _parse_month_header(header: str) -> date | None:
    """Parse 'May 2026', '31 May 2026', 'Sept 2025' → first-of-month date."""
    if not header or not isinstance(header, str):
        return None
    s = header.strip().rstrip(",")
    # Try day-month-year (balance sheet) first
    m = re.match(r"(?:(\d{1,2})\s+)?([A-Za-z]+)\s+(\d{4})", s)
    if not m:
        return None
    day_str, mon_str, year_str = m.groups()
    mon = _MONTH_NAMES.get(mon_str.lower())
    if not mon:
        return None
    year = int(year_str)
    day = int(day_str) if day_str else 1
    try:
        return date(year, mon, day)
    except ValueError:
        return None


def _read_xero_pl(path: Path) -> pd.DataFrame:
    """Parse Xero P&L into long-form: (account, account_lower, section, month, value)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Find the header row: contains 'Account' in col A and month headers in subsequent cols.
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.strip().lower() == "account":
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame(columns=["account", "account_lower", "section", "month", "value"])
    # Build month columns
    month_cols: list[tuple[int, date]] = []
    for c in range(2, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        d = _parse_month_header(h) if h else None
        if d:
            # Snap to first-of-month
            d = date(d.year, d.month, 1)
            month_cols.append((c, d))
    rows = []
    section = None
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        if not isinstance(a, str):
            continue
        a_clean = a.strip()
        if not a_clean:
            continue
        a_lower = a_clean.lower()
        # Detect section headers (no values in any month column, or known section labels)
        if a_lower in {"income", "revenue", "trading income", "cost of sales", "cost of goods sold section", "operating expenses", "less operating expenses"}:
            if a_lower in {"income", "revenue", "trading income"}:
                section = "Revenue"
            elif "cost of sales" in a_lower:
                section = "Cost of Sales"
            else:
                section = "Operating Expenses"
            continue
        # Skip totals + net profit lines
        if a_lower.startswith("total ") or a_lower in {"gross profit", "net profit", "net loss"}:
            continue
        # Capture per-month values
        for c, mdate in month_cols:
            v = ws.cell(r, c).value
            try:
                val = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            rows.append({
                "account": a_clean,
                "account_lower": a_lower,
                "section": section or "Operating Expenses",
                "month": mdate,
                "value": val,
            })
    return pd.DataFrame(rows)


def _read_xero_bs(path: Path) -> pd.DataFrame:
    """Parse Xero Balance Sheet → long-form (account, section, as_at, value)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Header row has 'Account' somewhere in col A or B
    header_row = None
    label_col = None
    for r in range(1, min(ws.max_row + 1, 20)):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "account":
                header_row = r
                label_col = c
                break
        if header_row:
            break
    if header_row is None:
        return pd.DataFrame(columns=["account", "section", "as_at", "value"])
    # Parse as-at dates from columns after label_col
    date_cols: list[tuple[int, date]] = []
    for c in range(label_col + 1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        d = _parse_month_header(h) if h else None
        if d:
            date_cols.append((c, d))
    _SUBSECTIONS = {"bank", "current assets", "fixed assets", "non-current assets",
                    "current liabilities", "non-current liabilities"}
    rows = []
    section = None
    subsection = None
    for r in range(header_row + 1, ws.max_row + 1):
        # Section header lives in col 1 when label_col == 2
        col1 = ws.cell(r, 1).value
        label = ws.cell(r, label_col).value
        if isinstance(col1, str) and col1.strip() and (label is None or label == ""):
            section_text = col1.strip()
            if section_text.lower() in {"assets", "liabilities", "equity"}:
                section = section_text
                subsection = None  # reset on a new top-level section
            continue
        if not isinstance(label, str) or not label.strip():
            continue
        label_clean = label.strip()
        label_lower = label_clean.lower()
        # Sub-section labels (e.g. 'Bank') carry the actual balances in their child
        # rows — Xero often exports the 'Total <sub-section>' line as 0, so we track
        # the sub-section and let downstream code sum the children instead.
        if label_lower in _SUBSECTIONS:
            subsection = label_clean
            continue
        # Skip totals + the 'Net Assets' summary line
        if label_lower.startswith("total "):
            continue
        if label_lower == "net assets":
            continue
        for c, d in date_cols:
            v = ws.cell(r, c).value
            try:
                val = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            rows.append({
                "account": label_clean,
                "section": section or "Assets",
                "subsection": subsection,
                "as_at": d,
                "value": val,
            })
    return pd.DataFrame(rows)


_ATXN_DEBIT_RE = re.compile(r"\bdebit\b", re.IGNORECASE)
_ATXN_CREDIT_RE = re.compile(r"\bcredit\b", re.IGNORECASE)


def _xero_atxn_column_map(ws, header_row: int) -> dict[str, int | None]:
    """Map logical fields to 1-based column indices from the header row."""
    headers: list[tuple[int, str]] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            headers.append((c, str(v).strip()))

    def first_match(pred) -> int | None:
        for c, h in headers:
            if pred(h):
                return c
        return None

    debit_candidates = [(c, h) for c, h in headers if _ATXN_DEBIT_RE.search(h)]
    credit_candidates = [(c, h) for c, h in headers if _ATXN_CREDIT_RE.search(h)]

    def pick_amount(candidates: list[tuple[int, str]]) -> int | None:
        if not candidates:
            return None
        # Multi-currency exports carry Debit(Source)/Credit(Source) plus reporting-ccy cols.
        non_source = [(c, h) for c, h in candidates if "source" not in h.lower()]
        pool = non_source or candidates
        return pool[-1][0]

    return {
        "date": first_match(lambda h: h.lower() == "date"),
        "contact": first_match(lambda h: h.lower().startswith("contact")),
        "description": first_match(lambda h: "description" in h.lower()),
        "source": first_match(lambda h: h.lower() == "source"),
        "reference": first_match(lambda h: "reference" in h.lower()),
        "currency": first_match(lambda h: "currency" in h.lower()),
        "debit": pick_amount(debit_candidates),
        "credit": pick_amount(credit_candidates),
        "gross": first_match(lambda h: h.lower().startswith("gross")),
        "gst": first_match(lambda h: h.lower().startswith("gst")),
    }


def _xero_atxn_cell(ws, row: int, col: int | None):
    if col is None:
        return None
    return ws.cell(row, col).value


def _xero_atxn_amount(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _read_xero_atxn(path: Path) -> pd.DataFrame:
    """Parse Xero Account Transactions into long-form (date, account, contact, source, description, debit, credit, gross, gst).

    Section-banded by account in column A. Column positions vary by export — debit/credit
    (and other fields) are resolved from header labels, not fixed positions.

    Supported layouts include:
      Date, Contact, Description, Debit (AUD), Credit (AUD)
      Date, Source, Contact, Description, Reference, Debit, Credit, ...
      Date, Source, Description, Reference, Currency, Debit(Source), Credit(Source),
      Debit(NZD), Credit(NZD), ...
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Find header row by locating the cell holding "Date" in column A.
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "date":
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame(columns=["date", "account", "contact", "source", "description", "debit", "credit", "gross", "gst"])

    cols = _xero_atxn_column_map(ws, header_row)
    if cols["debit"] is None or cols["credit"] is None:
        return pd.DataFrame(columns=["date", "account", "contact", "source", "description", "debit", "credit", "gross", "gst"])

    amount_cols = tuple(c for c in (cols["debit"], cols["credit"], cols["gross"], cols["gst"]) if c is not None)
    detail_cols = tuple(c for c in (
        cols["contact"], cols["description"], cols["source"], cols["reference"], cols["currency"],
    ) if c is not None)

    rows = []
    current_account: str | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        # Detect section header (str in col A, nothing meaningful in others)
        if isinstance(a, str) and not isinstance(a, (int, float)):
            a_str = a.strip()
            if a_str:
                a_lower = a_str.lower()
                if a_lower.startswith("opening balance") or a_lower.startswith("closing balance"):
                    continue
                if a_lower.startswith("total "):
                    continue
                non_blank = sum(
                    1 for c in detail_cols + amount_cols
                    if _xero_atxn_cell(ws, r, c) not in (None, "")
                )
                if non_blank == 0:
                    current_account = a_str
                    continue
        # Detect a real transaction row: Date in column A
        if isinstance(a, (datetime, date)):
            tx_date = a.date() if isinstance(a, datetime) else a
            desc_raw = _xero_atxn_cell(ws, r, cols["description"])
            desc_text = desc_raw.strip() if isinstance(desc_raw, str) else (str(desc_raw) if desc_raw is not None else "")
            contact_val = _xero_atxn_cell(ws, r, cols["contact"])
            if contact_val is not None and str(contact_val).strip():
                contact = str(contact_val).strip()
            elif " - " in desc_text:
                contact = desc_text.split(" - ", 1)[0].strip()
            else:
                contact = desc_text or "Unknown"
            currency_val = _xero_atxn_cell(ws, r, cols["currency"])
            rows.append({
                "date": tx_date,
                "account": current_account or "Unknown Account",
                "contact": contact,
                "source": _xero_atxn_cell(ws, r, cols["source"]) or "",
                "description": desc_text,
                "reference": _xero_atxn_cell(ws, r, cols["reference"]) or "",
                "currency": str(currency_val).strip() if currency_val else None,
                "debit": _xero_atxn_amount(_xero_atxn_cell(ws, r, cols["debit"])),
                "credit": _xero_atxn_amount(_xero_atxn_cell(ws, r, cols["credit"])),
                "gross": _xero_atxn_amount(_xero_atxn_cell(ws, r, cols["gross"])),
                "gst": _xero_atxn_amount(_xero_atxn_cell(ws, r, cols["gst"])),
            })
    return pd.DataFrame(rows)


# ---------- Entry point ----------


def ingest(inputs_dir: Path, run_date: date | None = None) -> IngestBundle:
    inputs_dir = Path(inputs_dir).resolve()
    if not inputs_dir.is_dir():
        raise FileNotFoundError(f"Inputs directory not found: {inputs_dir}")
    run_date = run_date or date.today()
    roles = _scan(inputs_dir)

    files_found = {k: str(v) for k, v in roles.items()}
    # xero_pl is optional — Account Transactions is the primary source and the P&L is
    # reconstructed from it, so its absence must not be reported as a missing input.
    expected = ["shopify_daily", "nc_rc", "sessions", "xero_bs", "xero_atxn", "ad_spend_meta"]
    missing = [r for r in expected if r not in roles]

    meta = IngestMeta(
        inputs_dir=str(inputs_dir),
        run_date=run_date,
        client_name=_client_name_from_files(roles),
        files_found=files_found,
        files_missing=missing,
    )

    bundle = IngestBundle(meta=meta)

    if "shopify_daily" in roles:
        bundle.shopify_daily = _read_shopify_daily(roles["shopify_daily"])
        if not bundle.shopify_daily.empty:
            meta.lookback_start = bundle.shopify_daily["day"].min()
            meta.lookback_end = bundle.shopify_daily["day"].max()
        # Shopify shop currency: unmarked in the export. Assume reporting_currency unless overridden.
        meta.shopify_currency = CurrencyTag(code=meta.reporting_currency, confidence="assumed",
                                            source="Shopify export does not declare currency in CSV; assumed shop currency = reporting currency")
    if "nc_rc" in roles:
        bundle.nc_rc = _read_nc_rc(roles["nc_rc"])
    if "sessions" in roles:
        bundle.sessions = _read_sessions(roles["sessions"])
    if "xero_pl" in roles:
        bundle.xero_pl = _read_xero_pl(roles["xero_pl"])
    if "xero_bs" in roles:
        bundle.xero_bs = _read_xero_bs(roles["xero_bs"])
    if "xero_atxn" in roles:
        bundle.xero_atxn = _read_xero_atxn(roles["xero_atxn"])
    # Ad spend
    for role_key, platform in (("ad_spend_meta", "meta"), ("ad_spend_google", "google"), ("ad_spend_tiktok", "tiktok")):
        if role_key in roles:
            df, ccy = _read_ad_spend(roles[role_key])
            bundle.ad_spend[platform] = df
            meta.ad_platform_currency[platform] = ccy
    if "cohort" in roles:
        bundle.cohort = _read_cohort(roles["cohort"])

    return bundle
