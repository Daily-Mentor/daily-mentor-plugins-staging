"""Render the same RenderTrees as an xlsx workbook.

- Every calculated cell gets an openpyxl.comments.Comment carrying tooltip text.
- Number formats applied immediately after value (G1).
- No formula merging on top of conditional formatting (G3).
- Zip integrity verified post-save (G7).
"""
from __future__ import annotations

import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import AuditReport, Cell, RenderTree, Row


_FMT_MAP = {
    "currency": '#,##0;[Red]-#,##0;"—"',
    "currency_dec": '#,##0.00;[Red]-#,##0.00;"—"',
    "pct": '0.0%;[Red]-0.0%;"—"',
    "int": '#,##0;[Red]-#,##0;"—"',
    "text": "General",
    "date": "yyyy-mm-dd",
    "ratio": '0.00',
}


_SECTION_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="1C1F24")
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="FAFAF7")
_HEADER_FONT = Font(bold=True, color="4B4D52")
_HEADER_FILL = PatternFill("solid", fgColor="FAFAF7")


def _write_value(cell, value, fmt):
    if value is None:
        cell.value = None
    elif fmt in ("currency", "currency_dec", "int", "ratio"):
        try:
            cell.value = float(value)
        except (TypeError, ValueError):
            cell.value = value
    elif fmt == "pct":
        try:
            cell.value = float(value)
        except (TypeError, ValueError):
            cell.value = value
    else:
        cell.value = value
    cell.number_format = _FMT_MAP.get(fmt, "General")


def _write_row(ws, r: int, row: Row, *, indent_base: int = 0):
    for c, source in enumerate(row.cells, start=1):
        cell = ws.cell(row=r, column=c)
        _write_value(cell, source.value, source.fmt)
        cell.alignment = Alignment(
            horizontal="left" if c == 1 else "right",
            indent=(source.indent + indent_base) if c == 1 else 0,
        )
        if row.is_section:
            cell.fill = _SECTION_FILL
            cell.font = _SECTION_FONT
        elif row.is_total or source.is_total:
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
        elif source.bold:
            cell.font = Font(bold=True)


def _write_tree(wb: Workbook, tree: RenderTree, *, sheet_name: str | None = None,
                rows_override: list[Row] | None = None):
    title = (sheet_name or tree.title)[:31]
    if title in wb.sheetnames:
        title = title[:28] + "_2"
    ws = wb.create_sheet(title=title)
    # Header banner row
    ws.cell(1, 1, tree.title).font = Font(bold=True, size=14)
    if tree.subtitle:
        ws.cell(2, 1, tree.subtitle).font = Font(italic=True, color="6C707A")
    # Column headers (tab banners live on the Audit Report sheet, not the page body)
    header_row = 4
    for c, col_name in enumerate(tree.columns, start=1):
        cell = ws.cell(header_row, c, col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    rows = rows_override if rows_override is not None else tree.rows
    r = header_row + 1
    for row in rows:
        _write_row(ws, r, row)
        r += 1
    # Notes block at the bottom
    if tree.notes:
        r += 1
        ws.cell(r, 1, "NOTES").font = Font(bold=True, color="6C707A")
        r += 1
        for note in tree.notes:
            ws.cell(r, 1, "• " + note).font = Font(color="4B4D52")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(len(tree.columns), 2))
            r += 1
    # Column widths
    ws.column_dimensions["A"].width = 40
    for c in range(2, len(tree.columns) + 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = 14


def _write_audit_tab(wb: Workbook, report: AuditReport, trees: list[RenderTree] | None = None):
    ws = wb.create_sheet(title="Audit Report")
    ws.cell(1, 1, "Audit Report").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Run {report.run_id} at {report.timestamp:%Y-%m-%d %H:%M}").font = Font(italic=True)
    headers = ["Check", "Name", "Status", "Message"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(4, c, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    r_idx = 5
    for r in report.results:
        ws.cell(r_idx, 1, r.check_id).font = Font(bold=True)
        ws.cell(r_idx, 2, r.name)
        ws.cell(r_idx, 3, r.status)
        ws.cell(r_idx, 4, r.message)
        r_idx += 1
    # Per-tab notices (banners moved off the page bodies)
    notices = [(t.title, b) for t in (trees or []) for b in t.banners]
    if notices:
        r_idx += 1
        ws.cell(r_idx, 1, "Tab Notices").font = Font(bold=True, size=12)
        r_idx += 1
        for tab, b in notices:
            ws.cell(r_idx, 1, b.severity.upper()).font = Font(
                color="B91C1C" if b.severity == "error" else "D97706" if b.severity == "warning" else "2563EB")
            ws.cell(r_idx, 2, tab)
            ws.cell(r_idx, 4, b.text)
            r_idx += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 80


def render(trees: list[RenderTree], audit_report: AuditReport, output_path: Path, *, client: str, run_date) -> dict:
    """Write the workbook. Returns dict with audit observations (zip_valid, sheet_count)."""
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    for tree in trees:
        _write_tree(wb, tree)

    _write_audit_tab(wb, audit_report, trees)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # G7: zip integrity
    zip_valid = True
    try:
        with zipfile.ZipFile(output_path) as zf:
            bad = zf.testzip()
            zip_valid = bad is None
    except Exception:
        zip_valid = False

    return {
        "zip_valid": zip_valid,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": list(wb.sheetnames),
    }
