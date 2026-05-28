"""Generate a synthetic, brand-neutral input pack for testing.

Produces a complete pack for a fictional brand ('Acme Goods Ltd') with GBP as the
shop/reporting currency and EUR ad spend — exercising a different brand name and a
non-AUD FX path. No real client data. Used by test_gates.py and runnable standalone:

    python3 -m tests.make_synthetic_pack /tmp/acme-inputs
"""
from __future__ import annotations

import csv
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook


CLIENT = "Acme Goods Ltd"
SHOP_CCY = "GBP"
AD_CCY = "EUR"


def _months_back(end: date, n: int) -> list[date]:
    out = []
    y, m = end.year, end.month
    for _ in range(n):
        out.append(date(y, m, 1))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(out))


def _daterange(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def write_pack(out_dir: Path, end: date | None = None) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    end = end or date(2026, 5, 20)
    start = end - timedelta(days=364)

    # ---- shopify_daily_sales_12mo.csv ----
    sd_path = out_dir / "shopify_daily_sales_12mo.csv"
    with sd_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Day", "Orders", "Gross sales", "Discounts", "Returns", "Net sales",
                    "Shipping charges", "Duties", "Additional fees", "Taxes", "Total sales",
                    "Cost of goods sold", "Net items sold"])
        for i, day in enumerate(_daterange(start, end)):
            base = 800 + (i % 30) * 25  # gentle variation
            orders = 5 + (i % 7)
            gross = round(base, 2)
            disc = round(-gross * 0.05, 2)
            ret = round(-gross * 0.01, 2)
            net = round(gross + disc + ret, 2)
            ship = 0
            tax = round(net * 0.10, 2)
            total = round(net + tax, 2)
            cogs = round(gross * 0.32, 2)
            units = orders + 1
            w.writerow([day.isoformat(), orders, gross, disc, ret, net, ship, 0, 0, tax, total, cogs, units])

    # ---- NC/RC ----
    ncrc_path = out_dir / "Gross sales by new or returning customer - 2026-05-20.csv"
    with ncrc_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["New or returning customer", "Gross sales", "Discounts", "Returns",
                    "Shipping charges", "Taxes", "Orders", "Average order value", "Cost of goods sold"])
        w.writerow(["New", 38000.0, -1900.0, -380.0, 0, 4180.0, 300, 126.33, 12160.0])
        w.writerow(["Returning", 9000.0, -450.0, -90.0, 0, 990.0, 80, 105.50, 2880.0])

    # ---- Sessions ----
    sess_path = out_dir / "Sessions by month - 2026-05-20.csv"
    with sess_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Month", "Online store visitors", "Sessions", "Conversion rate"])
        for m in _months_back(end, 13):
            visitors = 4000 + (m.month * 150)
            sessions = int(visitors * 1.2)
            w.writerow([m.isoformat(), visitors, sessions, 0.026])

    # ---- Ad spend (EUR) ----
    ad_path = out_dir / "facebook_spend.csv"
    with ad_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Day", "Campaign name", f"Amount spent ({AD_CCY})", "Reporting starts", "Reporting ends"])
        for i, day in enumerate(_daterange(start, end)):
            spend = round(300 + (i % 20) * 12, 2)
            w.writerow([day.isoformat(), "Prospecting | EU", spend, day.isoformat(), day.isoformat()])

    # ---- Xero P&L (sparse — forces Atxn-derived fallback) ----
    pl_path = out_dir / f"{CLIENT.replace(' ', '_')}_-_Profit_and_Loss.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Profit and Loss"
    ws["A1"] = "Profit and Loss"; ws["A2"] = CLIENT; ws["A3"] = "For the month ended 31 May 2026"
    months = _months_back(end, 13)
    ws["A5"] = "Account"
    for c, m in enumerate(reversed(months), start=2):
        ws.cell(5, c, m.strftime("%b %Y"))
    ws["A7"] = "Operating Expenses"
    ws["A8"] = "Subscriptions"; ws.cell(8, 2, 220.0)  # only most-recent month
    wb.save(pl_path)

    # ---- Xero Balance Sheet ----
    bs_path = out_dir / f"{CLIENT.replace(' ', '_')}_-_Balance_Sheet.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Balance Sheet"
    ws["A1"] = "Balance Sheet"; ws["A2"] = CLIENT; ws["A3"] = "As at 31 May 2026"
    ws["B5"] = "Account"
    for c, m in enumerate(reversed(months[-6:]), start=3):
        ws.cell(5, c, m.strftime("%d %b %Y"))
    ws["A7"] = "Assets"
    ws["B8"] = "Bank"; ws["B9"] = "Business Account"; ws.cell(9, 3, 18500.0)
    ws["B11"] = "Current Assets"; ws["B12"] = "Inventory"; ws.cell(12, 3, 42000.0)
    ws["A16"] = "Liabilities"; ws["B17"] = "Current Liabilities"
    ws["B18"] = "GST"; ws.cell(18, 3, 2400.0)
    ws["A27"] = "Equity"; ws["B28"] = "Current Year Earnings"; ws.cell(28, 3, 31000.0)
    wb.save(bs_path)

    # ---- Xero Account Transactions (12 months, layout A with Contact column) ----
    atxn_path = out_dir / f"{CLIENT.replace(' ', '_')}_-_Account_Transactions.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Account Transactions"
    ws["A1"] = "Account Transactions"; ws["A2"] = CLIENT
    ws["A3"] = "For the period 1 June 2025 to 31 May 2026"
    ws["A5"] = "Date"; ws["B5"] = "Source"; ws["C5"] = "Contact"; ws["D5"] = "Description"
    ws["E5"] = "Reference"; ws["F5"] = "Debit"; ws["G5"] = "Credit"; ws["H5"] = "Running Balance"
    ws["I5"] = "Gross"; ws["J5"] = "GST"
    r = 7
    sections = [
        ("Sales", [("Shopify Payments", 0, 4000.0)]),
        ("Cost of Goods Sold", [("Jun 25_Widget_CoGS Recognition", 1200.0, 0)]),
        ("Freight & Courier", [("Royal Mail", 320.0, 0), ("DPD", 180.0, 0)]),
        ("Subscriptions", [("Klaviyo", 90.0, 0), ("Shopify", 79.0, 0)]),
        ("Digital Marketing", [("Meta Ads", 5400.0, 0)]),
    ]
    for acct, txns in sections:
        ws.cell(r, 1, acct); r += 1
        for month in months[1:]:  # 12 months of postings
            for contact, debit, credit in txns:
                ws.cell(r, 1, month.replace(day=15))
                ws.cell(r, 2, "Spend Money" if debit else "Receive Money")
                ws.cell(r, 3, contact)
                ws.cell(r, 4, contact)
                ws.cell(r, 6, debit)
                ws.cell(r, 7, credit)
                r += 1
        r += 1
    wb.save(atxn_path)

    return out_dir


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/acme-inputs")
    p = write_pack(target)
    print(f"Synthetic pack for '{CLIENT}' ({SHOP_CCY} shop, {AD_CCY} ad spend) written to {p}")
