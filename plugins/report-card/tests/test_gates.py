"""Deployment acceptance gates for the report-card skill.

These three gates assert the skill is fit for marketplace distribution:

  Gate 1  Evergreen — works for an arbitrary brand and reporting currency.
  Gate 2  Pre-flight gates — blocks when required inputs are missing.
  Gate 3  Dual output — produces both HTML and xlsx.

Network-resilient: FX falls back gracefully if the rate API is unreachable,
so these run offline (conversions just use rate 1.0 in that case).

Run:  cd plugins/report-card && python3 -m pytest tests/test_gates.py -v
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts import check_deps as deps_mod
from scripts.cli import main as cli_main
from scripts.preflight import preflight
from tests.make_synthetic_pack import CLIENT, write_pack


# ---- Gate 0: Dependency gate ----

def test_gate0_deps_satisfied_in_this_env():
    report = deps_mod.check()
    assert report["ready"] is True
    assert report["python"]["ok"] is True


def test_gate0_reports_install_command_when_missing(monkeypatch):
    """Simulate a missing package and confirm the gate surfaces an install command."""
    real_find_spec = deps_mod.importlib.util.find_spec

    def fake_find_spec(name, *a, **k):
        if name == "openpyxl":
            return None
        return real_find_spec(name, *a, **k)

    monkeypatch.setattr(deps_mod.importlib.util, "find_spec", fake_find_spec)
    report = deps_mod.check()
    assert report["ready"] is False
    assert "openpyxl>=3.1" in report["missing_pip_names"]
    assert report["install_command"] and "pip install" in report["install_command"]
    # The human-readable text must name the package so the skill can offer to install it.
    assert "openpyxl" in deps_mod.render_text(report)


def test_gate0_no_baked_fx_dictionary_either():
    """Belt-and-braces: confirm the FX module is stdlib-only (no requests dependency)."""
    fx_src = (Path(__file__).resolve().parent.parent / "scripts" / "fx.py").read_text()
    assert "import requests" not in fx_src and "from requests" not in fx_src
    assert "urllib" in fx_src


@pytest.fixture(scope="module")
def synthetic_inputs(tmp_path_factory) -> Path:
    d = tmp_path_factory.mktemp("acme-inputs")
    write_pack(d, end=date(2026, 5, 20))
    return d


# ---- Gate 1: Evergreen ----

def test_gate1_evergreen_arbitrary_brand_and_currency(synthetic_inputs, tmp_path):
    out = tmp_path / "out"
    rc = cli_main([str(synthetic_inputs), str(out), "--reporting-currency", "GBP", "--run-date", "2026-05-28"])
    assert rc in (0, 1)  # 0 clean, 1 = non-halting audit FAIL (e.g. A4 reconciliation) — both are successful builds

    html = (out / "report-card-2026-05-28.html").read_text()
    # Brand-neutral: the synthetic client name flows through; no hardcoded brand.
    assert CLIENT in html
    assert "Boosh" not in html and "Sunday Mass" not in html
    # Reporting currency honoured.
    assert "GBP" in html


def test_gate1_no_baked_fx_dictionary():
    """There must be no static FX rate file shipped with the skill."""
    data_dir = Path(__file__).resolve().parent.parent / "data"
    assert not (data_dir / "fx_cache.json").exists(), "A baked-in FX dictionary must not ship — rates are fetched at runtime."


# ---- Gate 2: Pre-flight gates ----

def test_gate2_preflight_blocks_when_incomplete(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    report = preflight(empty)
    assert report.is_ready is False
    assert len(report.missing_required) == 7  # all required inputs absent


def test_gate2_preflight_ready_when_complete(synthetic_inputs):
    report = preflight(synthetic_inputs)
    assert report.is_ready is True
    assert report.missing_required == []
    # Optional inputs are surfaced but don't block.
    assert len(report.missing_optional) >= 1


def test_gate2_cli_refuses_build_when_incomplete(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    rc = cli_main([str(empty), str(tmp_path / "out")])
    assert rc == 2  # refuses to build
    assert not (tmp_path / "out").exists() or not list((tmp_path / "out").glob("report-card-*"))


# ---- Gate 3: Dual output ----

def test_gate3_produces_html_and_xlsx(synthetic_inputs, tmp_path):
    out = tmp_path / "out"
    cli_main([str(synthetic_inputs), str(out), "--reporting-currency", "GBP", "--run-date", "2026-05-28"])

    html_path = out / "report-card-2026-05-28.html"
    xlsx_path = out / "report-card-2026-05-28.xlsx"
    assert html_path.exists() and html_path.stat().st_size > 10_000
    assert xlsx_path.exists() and xlsx_path.stat().st_size > 10_000

    # xlsx opens, has the expected tab set, and carries no leftover cell comments.
    wb = load_workbook(xlsx_path)
    assert len(wb.sheetnames) >= 11
    assert "Monthly P&L" in wb.sheetnames
    assert "Audit Report" in wb.sheetnames
    comments = sum(1 for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row if c.comment)
    assert comments == 0, "xlsx must ship without cell comments (tooltips live in the HTML)."
