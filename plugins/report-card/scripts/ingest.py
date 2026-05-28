"""File discovery and parsing. No business logic — just typed frames.

The 7-file pack:
    shopify_daily_sales_*.csv
    Gross sales by new or returning customer - *.csv
    Sessions by month - *.csv
    *_Profit_and_Loss.xlsx           (Xero)
    *_Balance_Sheet.xlsx             (Xero)
    *_Account_Transactions.xlsx      (Xero)
    facebook_spend.{csv,xlsx}        (optional: google_spend, tiktok_spend)
    Optional: shopify cohort export

Currency is sniffed from CSV column headers (e.g. `Amount spent (NZD)`).
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .models import CurrencyTag, IngestBundle, IngestMeta


# ---------- File discovery ----------


def _scan(inputs_dir: Path) -> dict[str, Path]:
    """Map role → file path. Tolerant of naming variations.

    Multi-entity packs (e.g. parent + subsidiary Xero exports) — pick the
    largest file for each Xero role. Larger = more rows = more coverage =
    the operating entity. Other entities' files are dropped from the build.
    """
    files = {p.name: p for p in inputs_dir.iterdir() if p.is_file()}
    roles: dict[str, Path] = {}
    candidate_atxn: list[Path] = []
    candidate_pl: list[Path] = []
    candidate_bs: list[Path] = []
    for name, path in files.items():
        lower = name.lower()
        if "shopify_daily" in lower and lower.endswith(".csv"):
            roles["shopify_daily"] = path
        elif "gross sales by new or returning" in lower:
            roles["nc_rc"] = path
        elif "sessions by month" in lower:
            roles["sessions"] = path
        elif "profit_and_loss" in lower and lower.endswith(".xlsx"):
            candidate_pl.append(path)
        elif "balance_sheet" in lower and lower.endswith(".xlsx"):
            candidate_bs.append(path)
        elif "account_transactions" in lower and lower.endswith(".xlsx"):
            candidate_atxn.append(path)
        elif "facebook" in lower and "spend" in lower:
            roles.setdefault("ad_spend_meta", path)
        elif "google" in lower and "spend" in lower:
            roles["ad_spend_google"] = path
        elif "tiktok" in lower and "spend" in lower:
            roles["ad_spend_tiktok"] = path
        elif "cohort" in lower:
            roles["cohort"] = path
    if candidate_atxn:
        roles["xero_atxn"] = max(candidate_atxn, key=lambda p: p.stat().st_size)
    if candidate_pl:
        roles["xero_pl"] = max(candidate_pl, key=lambda p: p.stat().st_size)
    if candidate_bs:
        roles["xero_bs"] = max(candidate_bs, key=lambda p: p.stat().st_size)
    return roles


def _client_name_from_files(roles: dict[str, Path]) -> str:
    """Infer client name from a Xero export filename like '<Client>_-_Profit_and_Loss.xlsx'."""
    for k in ("xero_pl", "xero_bs", "xero_atxn"):
        if k in roles:
            stem = roles[k].stem  # e.g. "Boosh_PTY_LTD_-_Profit_and_Loss"
            # Strip the report suffix
            for suffix in ("_-_Profit_and_Loss", "_-_Balance_Sheet", "_-_Account_Transactions"):
                if stem.endswith(suffix):
                    return stem[: -len(suffix)].replace("_", " ")
    return "Unknown Client"


# ---------- CSV parsers ----------


def _read_shopify_daily(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # Normalise headers (some exports add leading/trailing whitespace, BOM)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "Day": "day",
        "Orders": "orders",
        "Gross sales": "gross",
        "Discounts": "discounts",
        "Returns": "returns",
        "Net sales": "net",
        "Shipping charges": "shipping",
        "Duties": "duties",
        "Additional fees": "additional_fees",
        "Taxes": "taxes",
        "Total sales": "total",
        "Cost of goods sold": "cogs",
        "Net items sold": "units",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "day" in df.columns:
        df["day"] = pd.to_datetime(df["day"], errors="coerce").dt.date
    for col in ("orders", "gross", "discounts", "returns", "net", "shipping", "taxes", "total", "cogs", "units"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.dropna(subset=["day"]).sort_values("day").reset_index(drop=True)


def _read_nc_rc(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "New or returning customer": "segment",
        "Gross sales": "gross",
        "Discounts": "discounts",
        "Returns": "returns",
        "Shipping charges": "shipping",
        "Taxes": "taxes",
        "Orders": "orders",
        "Average order value": "aov",
        "Cost of goods sold": "cogs",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    for col in ("gross", "discounts", "returns", "shipping", "taxes", "orders", "aov", "cogs"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def _read_sessions(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    rename = {
        "Month": "month",
        "Online store visitors": "visitors",
        "Sessions": "sessions",
        "Conversion rate": "cr",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "month" in df.columns:
        df["month"] = pd.to_datetime(df["month"], errors="coerce").dt.date
    for col in ("visitors", "sessions", "cr"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.dropna(subset=["month"]).sort_values("month").reset_index(drop=True)


_AD_CCY_RE = re.compile(r"Amount spent \(([A-Z]{3})\)")


def _read_ad_spend(path: Path) -> tuple[pd.DataFrame, CurrencyTag]:
    """Read FB/Google/TikTok spend CSV. Returns (df, currency_tag).

    Expected columns: Day, Campaign name, Amount spent (XXX), Reporting starts, Reporting ends.
    """
    df = pd.read_csv(path)
    df.columns = [c.strip().lstrip("﻿") for c in df.columns]
    # Find the spend column
    spend_col = None
    ccy = "AUD"
    ccy_confidence = "assumed"
    for col in df.columns:
        m = _AD_CCY_RE.match(col)
        if m:
            spend_col = col
            ccy = m.group(1)
            ccy_confidence = "explicit"
            break
    if spend_col is None:
        # Fallback: look for any 'Amount spent' / 'Spend' / 'Cost' column
        for col in df.columns:
            cl = col.lower()
            if "amount spent" in cl or cl == "spend" or cl == "cost":
                spend_col = col
                break
    if spend_col is None:
        raise ValueError(f"Could not find spend column in {path.name}")
    df = df.rename(columns={spend_col: "amount_orig"})
    rename = {"Day": "day", "Campaign name": "campaign", "Reporting starts": "start", "Reporting ends": "end"}
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "day" in df.columns:
        df["day"] = pd.to_datetime(df["day"], errors="coerce").dt.date
    df["amount_orig"] = pd.to_numeric(df["amount_orig"], errors="coerce").fillna(0)
    df["ccy"] = ccy
    df = df.dropna(subset=["day"]).reset_index(drop=True)
    return df, CurrencyTag(code=ccy, confidence=ccy_confidence, source=spend_col)


# ---------- Xero xlsx parsers ----------


_MONTH_NAMES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def _parse_month_header(header: str) -> date | None:
    """Parse 'May 2026', '31 May 2026', 'Sept 2025' → first-of-month date."""
    if not header or not isinstance(header, str):
        return None
    s = header.strip().rstrip(",")
    # Try day-month-year (balance sheet) first
    m = re.match(r"(?:(\d{1,2})\s+)?([A-Za-z]+)\s+(\d{4})", s)
    if not m:
        return None
    day_str, mon_str, year_str = m.groups()
    mon = _MONTH_NAMES.get(mon_str.lower())
    if not mon:
        return None
    year = int(year_str)
    day = int(day_str) if day_str else 1
    try:
        return date(year, mon, day)
    except ValueError:
        return None


def _read_xero_pl(path: Path) -> pd.DataFrame:
    """Parse Xero P&L into long-form: (account, account_lower, section, month, value)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Find the header row: contains 'Account' in col A and month headers in subsequent cols.
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.strip().lower() == "account":
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame(columns=["account", "account_lower", "section", "month", "value"])
    # Build month columns
    month_cols: list[tuple[int, date]] = []
    for c in range(2, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        d = _parse_month_header(h) if h else None
        if d:
            # Snap to first-of-month
            d = date(d.year, d.month, 1)
            month_cols.append((c, d))
    rows = []
    section = None
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        if not isinstance(a, str):
            continue
        a_clean = a.strip()
        if not a_clean:
            continue
        a_lower = a_clean.lower()
        # Detect section headers (no values in any month column, or known section labels)
        if a_lower in {"income", "revenue", "trading income", "cost of sales", "cost of goods sold section", "operating expenses", "less operating expenses"}:
            if a_lower in {"income", "revenue", "trading income"}:
                section = "Revenue"
            elif "cost of sales" in a_lower:
                section = "Cost of Sales"
            else:
                section = "Operating Expenses"
            continue
        # Skip totals + net profit lines
        if a_lower.startswith("total ") or a_lower in {"gross profit", "net profit", "net loss"}:
            continue
        # Capture per-month values
        for c, mdate in month_cols:
            v = ws.cell(r, c).value
            try:
                val = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            rows.append({
                "account": a_clean,
                "account_lower": a_lower,
                "section": section or "Operating Expenses",
                "month": mdate,
                "value": val,
            })
    return pd.DataFrame(rows)


def _read_xero_bs(path: Path) -> pd.DataFrame:
    """Parse Xero Balance Sheet → long-form (account, section, as_at, value)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Header row has 'Account' somewhere in col A or B
    header_row = None
    label_col = None
    for r in range(1, min(ws.max_row + 1, 20)):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "account":
                header_row = r
                label_col = c
                break
        if header_row:
            break
    if header_row is None:
        return pd.DataFrame(columns=["account", "section", "as_at", "value"])
    # Parse as-at dates from columns after label_col
    date_cols: list[tuple[int, date]] = []
    for c in range(label_col + 1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        d = _parse_month_header(h) if h else None
        if d:
            date_cols.append((c, d))
    rows = []
    section = None
    for r in range(header_row + 1, ws.max_row + 1):
        # Section header lives in col 1 when label_col == 2
        col1 = ws.cell(r, 1).value
        label = ws.cell(r, label_col).value
        if isinstance(col1, str) and col1.strip() and (label is None or label == ""):
            section_text = col1.strip()
            if section_text.lower() in {"assets", "liabilities", "equity"}:
                section = section_text
            continue
        if not isinstance(label, str) or not label.strip():
            continue
        label_clean = label.strip()
        label_lower = label_clean.lower()
        # Skip totals + sub-section labels
        if label_lower.startswith("total "):
            continue
        if label_lower in {"net assets", "bank", "current assets", "fixed assets", "non-current assets", "current liabilities", "non-current liabilities"}:
            continue
        for c, d in date_cols:
            v = ws.cell(r, c).value
            try:
                val = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                val = 0.0
            rows.append({
                "account": label_clean,
                "section": section or "Assets",
                "as_at": d,
                "value": val,
            })
    return pd.DataFrame(rows)


def _read_xero_atxn(path: Path) -> pd.DataFrame:
    """Parse Xero Account Transactions into long-form (date, account, contact, source, description, debit, credit, gross, gst).

    Section-banded by account in column A. Two known column layouts:

    Layout A (PTY LTD / single-currency):
        A:Date B:Source C:Contact D:Description E:Reference F:Debit G:Credit H:RunBal I:Gross J:GST

    Layout B (Limited / multi-currency, NZD-base):
        A:Date B:Source C:Description D:Reference E:Currency F:Debit(Source)
        G:Credit(Source) H:Debit(NZD) I:Credit(NZD) J:Running Balance(NZD)

    In Layout B the Contact is encoded inside Description as `"<Vendor> - <Memo>"`.
    We always emit cols (debit, credit) in the entity's reporting currency:
    Layout A -> F/G (single currency), Layout B -> H/I (NZD).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Find header row by locating the cell holding "Date" in column A.
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "date":
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame(columns=["date", "account", "contact", "source", "description", "debit", "credit", "gross", "gst"])

    # Detect layout by inspecting header cells
    header_c = (ws.cell(header_row, 3).value or "")
    header_e = (ws.cell(header_row, 5).value or "")
    layout_b = (
        "currency" in str(header_e).lower()
        or "description" in str(header_c).lower()
    )

    rows = []
    current_account: str | None = None
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        g = ws.cell(r, 7).value
        h = ws.cell(r, 8).value
        i = ws.cell(r, 9).value
        j = ws.cell(r, 10).value
        # Detect section header (str in col A, nothing meaningful in others)
        if isinstance(a, str) and not isinstance(a, (int, float)):
            a_str = a.strip()
            if a_str:
                a_lower = a_str.lower()
                if a_lower.startswith("opening balance") or a_lower.startswith("closing balance"):
                    continue
                if a_lower.startswith("total "):
                    continue
                non_blank = sum(1 for x in (b, c, d, e, f, g) if x not in (None, ""))
                if non_blank == 0:
                    current_account = a_str
                    continue
        # Detect a real transaction row: Date in column A
        if isinstance(a, (datetime, date)):
            tx_date = a.date() if isinstance(a, datetime) else a
            if layout_b:
                # Layout B (multi-currency Xero exports: Source ccy + reporting-ccy columns)
                desc_raw = c.strip() if isinstance(c, str) else (str(c) if c is not None else "")
                # Vendor extraction: "Vendor - Memo" → vendor; bare strings → whole string.
                if " - " in desc_raw:
                    vendor = desc_raw.split(" - ", 1)[0].strip()
                else:
                    vendor = desc_raw or "Unknown"
                debit_val = h if h not in (None, "") else 0.0
                credit_val = i if i not in (None, "") else 0.0
                src_ccy = (str(e).strip() if e else None)
                rows.append({
                    "date": tx_date,
                    "account": current_account or "Unknown Account",
                    "contact": vendor,
                    "source": b if b is not None else "",
                    "description": desc_raw,
                    "reference": d if d is not None else "",
                    "currency": src_ccy,
                    "debit": float(debit_val) if debit_val not in (None, "") else 0.0,
                    "credit": float(credit_val) if credit_val not in (None, "") else 0.0,
                    "gross": 0.0,
                    "gst": 0.0,
                })
            else:
                # Layout A (single-currency Xero exports with explicit Contact column)
                rows.append({
                    "date": tx_date,
                    "account": current_account or "Unknown Account",
                    "contact": (c.strip() if isinstance(c, str) else (c if c is not None else "Unknown")),
                    "source": b if b is not None else "",
                    "description": d if d is not None else "",
                    "reference": e if e is not None else "",
                    "currency": None,
                    "debit": float(f) if f not in (None, "") else 0.0,
                    "credit": float(g) if g not in (None, "") else 0.0,
                    "gross": float(i) if i not in (None, "") else 0.0,
                    "gst": float(j) if j not in (None, "") else 0.0,
                })
    return pd.DataFrame(rows)


# ---------- Entry point ----------


def ingest(inputs_dir: Path, run_date: date | None = None) -> IngestBundle:
    inputs_dir = Path(inputs_dir).resolve()
    if not inputs_dir.is_dir():
        raise FileNotFoundError(f"Inputs directory not found: {inputs_dir}")
    run_date = run_date or date.today()
    roles = _scan(inputs_dir)

    files_found = {k: str(v) for k, v in roles.items()}
    expected = ["shopify_daily", "nc_rc", "sessions", "xero_pl", "xero_bs", "xero_atxn", "ad_spend_meta"]
    missing = [r for r in expected if r not in roles]

    meta = IngestMeta(
        inputs_dir=str(inputs_dir),
        run_date=run_date,
        client_name=_client_name_from_files(roles),
        files_found=files_found,
        files_missing=missing,
    )

    bundle = IngestBundle(meta=meta)

    if "shopify_daily" in roles:
        bundle.shopify_daily = _read_shopify_daily(roles["shopify_daily"])
        if not bundle.shopify_daily.empty:
            meta.lookback_start = bundle.shopify_daily["day"].min()
            meta.lookback_end = bundle.shopify_daily["day"].max()
        # Shopify shop currency: unmarked in the export. Assume reporting_currency unless overridden.
        meta.shopify_currency = CurrencyTag(code=meta.reporting_currency, confidence="assumed",
                                            source="Shopify export does not declare currency in CSV; assumed shop currency = reporting currency")
    if "nc_rc" in roles:
        bundle.nc_rc = _read_nc_rc(roles["nc_rc"])
    if "sessions" in roles:
        bundle.sessions = _read_sessions(roles["sessions"])
    if "xero_pl" in roles:
        bundle.xero_pl = _read_xero_pl(roles["xero_pl"])
    if "xero_bs" in roles:
        bundle.xero_bs = _read_xero_bs(roles["xero_bs"])
    if "xero_atxn" in roles:
        bundle.xero_atxn = _read_xero_atxn(roles["xero_atxn"])
    # Ad spend
    for role_key, platform in (("ad_spend_meta", "meta"), ("ad_spend_google", "google"), ("ad_spend_tiktok", "tiktok")):
        if role_key in roles:
            df, ccy = _read_ad_spend(roles[role_key])
            bundle.ad_spend[platform] = df
            meta.ad_platform_currency[platform] = ccy
    if "cohort" in roles:
        try:
            bundle.cohort = pd.read_csv(roles["cohort"])
        except Exception:
            bundle.cohort = None

    return bundle
